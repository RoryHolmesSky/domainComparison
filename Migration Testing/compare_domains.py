from typing import overload
from pandas.core.algorithms import diff
import pandas as pd
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
import os
from pandas.core.indexes.base import Index
 

def add_domain_level_sheet(old, new, dom):
    old_instance = pd.read_excel(old, engine = 'openpyxl')
    new_instance = pd.read_excel(new, engine = 'openpyxl')
    for col in range(len(old_instance.columns)):
        if old_instance.dtypes[col] == "object":
            try:
                old_instance[old_instance.columns[col]] = old_instance[old_instance.columns[col]].str.strip()
            except:
                continue
    for col in range(len(new_instance.columns)):
        if new_instance.dtypes[col] == "object":
            try:
                new_instance[new_instance.columns[col]] = new_instance[new_instance.columns[col]].str.strip()
            except:
                continue
    if dom == "ARD":
        old_instance = old_instance.drop(['Roles > Owner > User Name', 'Roles > Owner > First Name', 'Roles > Owner > Last Name', 'Roles > Owner > Group Name',
                                          'Roles > Technical Steward > User Name', 'Roles > Technical Steward > First Name', 'Roles > Technical Steward > Last Name', 'Roles > Technical Steward > Group Name', 
                                          'Roles > Stakeholder > User Name', 'Roles > Stakeholder > First Name', 'Roles > Stakeholder > Last Name', 'Roles > Stakeholder > Group Name'], axis=1)
        new_instance = new_instance.drop(['Roles > Owner > User Name', 'Roles > Owner > First Name', 'Roles > Owner > Last Name', 'Roles > Owner > Group Name',
                                          'Roles > Technical Steward > User Name', 'Roles > Technical Steward > First Name', 'Roles > Technical Steward > Last Name', 'Roles > Technical Steward > Group Name', 
                                          'Roles > Stakeholder > User Name', 'Roles > Stakeholder > First Name', 'Roles > Stakeholder > Last Name', 'Roles > Stakeholder > Group Name'], axis=1)
    old_instance1 = old_instance.drop_duplicates()
    new_instance1 = new_instance.drop_duplicates()
    equal_check = new_instance1.equals(old_instance1)
    print (equal_check)
    if equal_check != True:
        columns = []
        all_columns = (old_instance.columns[:-1])
        for col in all_columns:
            columns.append(col)
        old_instance = old_instance.sort_values(by=columns)
        columns = []
        all_columns = (new_instance.columns[:-1])
        for col in all_columns:
            columns.append(col)
        new_instance = new_instance.sort_values(by=columns)
        df_merged = old_instance.merge(new_instance, how = 'outer' ,indicator=True)
        numTested = len(df_merged)
        df_merged['_merge'] = df_merged['_merge'].replace(['both','left_only','right_only'],['in both instances', 'old_instance_only', 'new_instance_only'])
        df_merged = df_merged.rename(columns={'_merge':'Collibra_Instance'})
        df_merged = df_merged[(df_merged.Collibra_Instance != "in both instances")]
        if df_merged.empty == False:
            columns = []
            all_columns = (df_merged.columns[:-1])
            for col in all_columns:
                columns.append(col)
            df_merged = df_merged.sort_values(by=columns)
            book = openpyxl.load_workbook('Migration Testing/domain_differences.xlsx')
            if dom in book.sheetnames:
                print ("sheet deleted")
                book.remove(book[dom])
            book.save('Migration Testing/domain_differences.xlsx')
            book.close()
            book = openpyxl.load_workbook('Migration Testing/domain_differences.xlsx')
            writer = pd.ExcelWriter('Migration Testing/domain_differences.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_merged.to_excel(writer, sheet_name=dom, index = False)
            writer.save()
            writer.close()
            num_of_pairs = int(len(df_merged)/2)
            if num_of_pairs == 0:
                width = df_merged.shape[1]
                for wid in range(width):
                    excel_path = openpyxl.load_workbook('C:/Users/RHS18/Documents/gitHubSky/domainComparison/Migration Testing/domain_differences.xlsx')
                    currentSheet1 = excel_path[dom]
                    redFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
                    currentSheet1.cell(row=2, column=wid+1).fill = redFill
                    excel_path.save(filename='C:/Users/RHS18/Documents/gitHubSky/domainComparison/Migration Testing/domain_differences.xlsx')
                    result = "FAIL"
                    rowsFailed = 1
                    numSchemas = 0
                    numTables = 0
                    schemaNames = []
                    tableNames = []
                    for i in range(len(df_merged)):
                        if df_merged['Asset Type'].iloc[i] == 'Schema':
                            schemaNames.append(df_merged['Name'].iloc[i])
                            numSchemas += 1
                        if df_merged['Asset Type'].iloc[i] == 'Table':
                            tableNames.append(df_merged['Name'].iloc[i])
                            numTables += 1       
                    if numSchemas == 0:
                        schemaNames = pd.NA
                        numSchemas = pd.NA
                    else:
                        schemaNames = str(schemaNames).replace("[", "").replace("]","")
                    if numTables == 0:
                        tableNames = pd.NA
                        numTables = pd.NA
                    else:
                        tableNames = str(tableNames).replace("[", "").replace("]","")
            else:
                print (num_of_pairs)
                excel_path = openpyxl.load_workbook('C:/Users/RHS18/Documents/gitHubSky/domainComparison/Migration Testing/domain_differences.xlsx')
                redFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
                for pair in range(num_of_pairs):
                    print (pair)
                    pair = pair + 1
                    df_zipped = zip(df_merged.values[(pair*2)-2], df_merged.values[(pair*2)-1])
                    df_zipped = (list(df_zipped))[:-1]
                    for each in range(len(df_zipped)):
                        if df_zipped[each][0] != df_zipped[each][1]:
                            if (pd.isna(df_zipped[each][0]) == False) & (pd.isna(df_zipped[each][0]) == False):
                                currentSheet1 = excel_path[dom]
                                currentSheet1.cell(row=(pair*2)+1, column=each+1).fill = redFill
                                currentSheet1.cell(row=(pair*2), column=each+1).fill = redFill
                                result = "FAIL"
                                rowsFailed = len(df_merged) / 2
                                numSchemas = 0
                                numTables = 0
                                schemaNames = []
                                tableNames = []
                    try:
                        for i in range(len(df_merged)):
                            if df_merged['Asset Type'].iloc[i] == 'Schema':
                                schemaNames.append(df_merged['Name'].iloc[i])
                                numSchemas += 1
                            if df_merged['Asset Type'].iloc[i] == 'Table':
                                tableNames.append(df_merged['Name'].iloc[i])
                                numTables += 1       
                        if numSchemas == 0:
                            schemaNames = pd.NA
                            numSchemas = pd.NA
                        else:
                            schemaNames = str(schemaNames).replace("[", "").replace("]","")
                        if numTables == 0:
                            tableNames = pd.NA
                            numTables = pd.NA
                        else:
                            tableNames = str(tableNames).replace("[", "").replace("]","")
                    except:
                        continue
                excel_path.save(filename='C:/Users/RHS18/Documents/gitHubSky/domainComparison/Migration Testing/domain_differences.xlsx')
        else:
            equals = {'Domain': ['Old Instance', 'New Instance'],
            'Result': ['Equal', 'Equal']
            }
            equalDF = pd.DataFrame(equals, columns = ['Domain', 'Result'])
            book = openpyxl.load_workbook('Migration Testing/domain_differences.xlsx')
            if dom in book.sheetnames:
                print ("sheet deleted")
                book.remove(book[dom])
            writer = pd.ExcelWriter('Migration Testing/domain_differences.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            equalDF.to_excel(writer, sheet_name=dom, index = False)
            writer.save()
            writer.close()
            result = "PASS"
            rowsFailed = pd.NA
            numSchemas = pd.NA
            numTables = pd.NA
            schemaNames = pd.NA
            tableNames = pd.NA
    else:
        equals = {'Domain': ['Old Instance', 'New Instance'],
        'Result': ['Equal', 'Equal']
        }
        equalDF = pd.DataFrame(equals, columns = ['Domain', 'Result'])
        book = openpyxl.load_workbook('Migration Testing/domain_differences.xlsx')
        if dom in book.sheetnames:
            print ("sheet deleted")
            book.remove(book[dom])
        writer = pd.ExcelWriter('Migration Testing/domain_differences.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        equalDF.to_excel(writer, sheet_name=dom, index = False)
        writer.save()
        writer.close()
        result = "PASS"
        rowsFailed = pd.NA
        numTested = len(old_instance)
        numSchemas = pd.NA
        numTables = pd.NA
        schemaNames = pd.NA
        tableNames = pd.NA
    print (result, numTested, rowsFailed, numSchemas, numTables, schemaNames, tableNames)
    return result, numTested, rowsFailed, numSchemas, numTables, schemaNames, tableNames

path = r"C:/Users/RHS18/Documents/gitHubSky/domainComparison/Migration Testing/Inputs"
domains = []
for files in os.listdir(path):
    domains.append(files.split(" - ")[0])
    file = (os.path.join(path, files)).replace("\\","/")
domains = list(set(domains))
print (domains)
result = []
numTested = []
rowsFailed = []
numSchemas = []
numTables = []
schemaNames = []
tableNames = []
for dom in domains:
    domainFiles = []
    for files in os.listdir(path):
        if files.startswith(dom):
            file = (os.path.join(path, files)).replace("\\","/")
            domainFiles.append(file)
    data = add_domain_level_sheet(domainFiles[1], domainFiles[0], dom)
    result.append(data[0])
    numTested.append(data[1])
    rowsFailed.append(data[2])
    numSchemas.append(data[3])
    numTables.append(data[4])
    schemaNames.append(data[5])
    tableNames.append(data[6])
summaryReport = {'Domain': domains,
    'Result': result,
    'Number of rows tested' : numTested,
    'Rows Failed' : rowsFailed,
    'New Schemas?' : numSchemas,
    'New Tables?' : numTables,
    'New Schema Names' : schemaNames,
    'New Table Names' : tableNames
    }
equalDF = pd.DataFrame(summaryReport, columns = ['Domain', 'Result', 'Number of rows tested', 'Rows Failed', 'New Schemas?', 'New Tables?', 'New Schema Names', 'New Table Names'])
print (equalDF)
width = equalDF.shape[1]
book = openpyxl.load_workbook('Migration Testing/domain_differences.xlsx')
redFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
greenFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='0000FF00'))
currentSheet1 = book['Summary']
print (len(equalDF))
for res in range(len(equalDF)):
    if equalDF['Result'].iloc[res] == 'FAIL':
        for wid in range(width):
            currentSheet1.cell(row=res+2, column=wid+1).fill = redFill
    elif equalDF['Result'].iloc[res] == 'PASS':
        for wid in range(width):
            currentSheet1.cell(row=res+2, column=wid+1).fill = greenFill
writer = pd.ExcelWriter('Migration Testing/domain_differences.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
equalDF.to_excel(writer, sheet_name= "Summary", index = False)
writer.save()
writer.close()
